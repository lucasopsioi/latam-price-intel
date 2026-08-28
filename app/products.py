# -*- coding: utf-8 -*-
"""我方产品主数据：Excel 模板生成、导入、CRUD。

字段按用户手绘草图①：
  产品营销名 / 产品内部代号 / 入网认证型号 / 产品颜色(SKU) / SKU EAN code /
  产品配置(RAM+ROM) / 其它配置(芯片/屏幕/是否旗舰屏) / 包装内清单 /
  产品ID图 / 最早发货时间 / 最晚发货时间 / 分国定价(RRP本币+汇率+USD+折扣)

导入策略：三张 sheet（产品 / SKU / 分国定价），用「产品营销名」当外键关联。
比让用户在一张宽表里填几十列友好得多，也避免同一产品重复填基础信息。
"""
from __future__ import annotations

import json
import logging
from pathlib import Path

from . import db

log = logging.getLogger("products")

SHEET_PRODUCT = "产品"
SHEET_SKU = "SKU"
SHEET_PRICING = "分国定价"

PRODUCT_COLS = [
    ("marketing_name", "产品营销名*", "ACME Vega 80 Pro"),
    ("internal_code", "产品内部代号", "PUR-AL80"),
    ("cert_model", "入网认证型号", "PUR-AL80"),
    ("category_code", "产业*(phone/wearable/audio/tablet/pc)", "phone"),
    ("series", "产品系列", "Vega"),
    ("chipset", "芯片", "Kirin 9020"),
    ("screen", "屏幕", "6.8英寸 LTPO OLED"),
    ("is_flagship_screen", "是否旗舰屏(1/0)", "1"),
    ("box_contents", "包装内清单(逗号分隔)", "充电器,数据线,保护壳,QSG,电子保卡"),
    ("launch_earliest", "最早发货时间(YYYY-MM-DD)", "2026-06-15"),
    ("launch_latest", "最晚发货时间(YYYY-MM-DD)", "2026-07-01"),
    ("status", "状态(active/eol/planned)", "active"),
    ("notes", "备注", ""),
]

SKU_COLS = [
    ("marketing_name", "产品营销名*", "ACME Vega 80 Pro"),
    ("sku_name", "SKU名", "SKU1"),
    ("color", "颜色", "羽砂黑"),
    ("ram_gb", "RAM(GB)", "12"),
    ("rom_gb", "ROM(GB)", "512"),
    ("ean_code", "EAN code", "6941487290123"),
    ("is_flagship_screen", "该SKU是否旗舰屏(1/0)", "1"),
]

PRICING_COLS = [
    ("marketing_name", "产品营销名*", "ACME Vega 80 Pro"),
    ("country_code", "国家*(MX/BR/CO/CL/PE/AR)", "MX"),
    ("rrp_local", "定价本币RRP*", "24999"),
    ("fx_rate", "定价汇率(1USD=?本币)", "18.5"),
    ("street_price_local", "实际街价", "23499"),
    ("discount_pct", "常促折扣%", "6"),
    ("on_sale", "是否在售(1/0)*", "1"),
    ("sales_volume_band", "销量档(大致即可)", "5k-10k/月"),
    ("sales_period", "销量对应周期", "2026-06~2026-08"),
]


# ---------------------------------------------------------------- 模板

def write_template(path: str | Path) -> Path:
    """生成带表头、说明行和示例行的 Excel 模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    path = Path(path)
    wb = Workbook()
    head_fill = PatternFill("solid", fgColor="1D6F42")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    note_font = Font(color="808080", italic=True, size=9)

    for idx, (sheet_name, cols, note) in enumerate([
        (SHEET_PRODUCT, PRODUCT_COLS,
         "一行一个产品。带*为必填。产业代码必须是 phone/wearable/audio/tablet/pc 之一。"),
        (SHEET_SKU, SKU_COLS,
         "一行一个 SKU（颜色×内存组合）。用「产品营销名」关联上一张表，必须完全一致。"),
        (SHEET_PRICING, PRICING_COLS,
         "一行 = 一个产品在一个国家的定价。同一产品在 6 个国家就填 6 行。"
         "销量给大致区间即可，不需要精确数字。"),
    ]):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet_name
        ws.cell(1, 1, note).font = note_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        for c, (_key, label, sample) in enumerate(cols, 1):
            h = ws.cell(2, c, label)
            h.fill, h.font = head_fill, head_font
            h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(3, c, sample).font = note_font
            ws.column_dimensions[h.column_letter].width = max(14, min(len(label) + 6, 34))
        ws.freeze_panes = "A3"

    wb.save(path)
    return path


# ---------------------------------------------------------------- 导入

def import_workbook(path: str | Path) -> dict:
    """导入 Excel。返回统计与逐条错误，不静默吞错。"""
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, data_only=True)
    report = {"products": 0, "skus": 0, "pricing": 0, "errors": [], "warnings": []}

    valid_cats = {r["code"] for r in db.q("SELECT code FROM category")}
    valid_countries = {r["code"] for r in db.q("SELECT code FROM country")}

    # —— 产品 ——
    name_to_id: dict[str, int] = {}
    for row in _rows(wb, SHEET_PRODUCT, PRODUCT_COLS, report):
        name = (row.get("marketing_name") or "").strip()
        if not name:
            continue
        cat = (row.get("category_code") or "").strip().lower()
        if cat not in valid_cats:
            report["errors"].append(
                f"[{SHEET_PRODUCT}] 「{name}」的产业代码「{cat}」无效，"
                f"必须是 {sorted(valid_cats)} 之一，该行已跳过")
            continue
        box = [x.strip() for x in str(row.get("box_contents") or "").split(",") if x.strip()]
        with db.tx() as conn:
            cur = conn.execute("""
                INSERT INTO my_product(marketing_name,internal_code,cert_model,
                    category_code,series,chipset,screen,is_flagship_screen,
                    box_contents,launch_earliest,launch_latest,status,notes,updated_at)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))
                ON CONFLICT(marketing_name, COALESCE(internal_code,'')) DO NOTHING
            """, (name, row.get("internal_code"), row.get("cert_model"), cat,
                  row.get("series"), row.get("chipset"), row.get("screen"),
                  _int(row.get("is_flagship_screen")) or 0,
                  json.dumps(box, ensure_ascii=False),
                  _date(row.get("launch_earliest")), _date(row.get("launch_latest")),
                  (row.get("status") or "active").strip(), row.get("notes")))
            pid = cur.lastrowid
        if not pid:
            exist = db.q1("SELECT id FROM my_product WHERE marketing_name=?", (name,))
            pid = exist["id"] if exist else None
            if pid:
                with db.tx() as conn:
                    conn.execute("""UPDATE my_product SET internal_code=?,cert_model=?,
                        category_code=?,series=?,chipset=?,screen=?,
                        box_contents=?,launch_earliest=?,launch_latest=?,
                        status=?,notes=?,updated_at=datetime('now') WHERE id=?""",
                                 (row.get("internal_code"), row.get("cert_model"), cat,
                                  row.get("series"), row.get("chipset"), row.get("screen"),
                                  json.dumps(box, ensure_ascii=False),
                                  _date(row.get("launch_earliest")),
                                  _date(row.get("launch_latest")),
                                  (row.get("status") or "active").strip(),
                                  row.get("notes"), pid))
        if pid:
            name_to_id[name] = pid
            report["products"] += 1

    # 补齐库里已有的产品名（支持只更新 SKU/定价 的增量导入）
    for r in db.q("SELECT id, marketing_name FROM my_product"):
        name_to_id.setdefault(r["marketing_name"], r["id"])

    # —— SKU ——
    for row in _rows(wb, SHEET_SKU, SKU_COLS, report):
        name = (row.get("marketing_name") or "").strip()
        pid = name_to_id.get(name)
        if not pid:
            if name:
                report["errors"].append(
                    f"[{SHEET_SKU}] 找不到产品「{name}」—— "
                    f"请确认与「{SHEET_PRODUCT}」表里的营销名完全一致")
            continue
        with db.tx() as conn:
            conn.execute("""
                INSERT INTO my_sku(product_id,sku_name,color,ram_gb,rom_gb,
                                   ean_code,is_flagship_screen)
                VALUES(?,?,?,?,?,?,?)
                ON CONFLICT(product_id, COALESCE(color,''), COALESCE(ram_gb,-1),
                            COALESCE(rom_gb,-1)) DO UPDATE SET
                  sku_name=excluded.sku_name, ean_code=excluded.ean_code,
                  is_flagship_screen=excluded.is_flagship_screen
            """, (pid, row.get("sku_name"), row.get("color"),
                  _int(row.get("ram_gb")), _int(row.get("rom_gb")),
                  str(row.get("ean_code") or "").strip() or None,
                  _int(row.get("is_flagship_screen"))))
        report["skus"] += 1

    # —— 分国定价 ——
    for row in _rows(wb, SHEET_PRICING, PRICING_COLS, report):
        name = (row.get("marketing_name") or "").strip()
        pid = name_to_id.get(name)
        cc = (row.get("country_code") or "").strip().upper()
        if not pid:
            if name:
                report["errors"].append(f"[{SHEET_PRICING}] 找不到产品「{name}」")
            continue
        if cc not in valid_countries:
            report["errors"].append(
                f"[{SHEET_PRICING}] 「{name}」的国家码「{cc}」无效，"
                f"必须是 {sorted(valid_countries)} 之一")
            continue
        rrp = _float(row.get("rrp_local"))
        if rrp is None:
            report["warnings"].append(f"[{SHEET_PRICING}] 「{name}/{cc}」没填定价，"
                                      f"竞品匹配会跳过这个国家")
        fx = _float(row.get("fx_rate"))
        currency = db.q1("SELECT currency FROM country WHERE code=?", (cc,))["currency"]
        with db.tx() as conn:
            conn.execute("""
                INSERT INTO my_pricing(product_id,sku_id,country_code,rrp_local,currency,
                    fx_rate,rrp_usd,street_price_local,discount_pct,on_sale,
                    sales_volume_band,sales_period,updated_at)
                VALUES(?,NULL,?,?,?,?,?,?,?,?,?,?,datetime('now'))
                ON CONFLICT(product_id, COALESCE(sku_id,0), country_code) DO UPDATE SET
                  rrp_local=excluded.rrp_local, fx_rate=excluded.fx_rate,
                  rrp_usd=excluded.rrp_usd, street_price_local=excluded.street_price_local,
                  discount_pct=excluded.discount_pct, on_sale=excluded.on_sale,
                  sales_volume_band=excluded.sales_volume_band,
                  sales_period=excluded.sales_period, updated_at=datetime('now')
            """, (pid, cc, rrp, currency, fx,
                  (rrp / fx) if (rrp and fx) else None,
                  _float(row.get("street_price_local")),
                  _float(row.get("discount_pct")),
                  1 if str(row.get("on_sale") or "1").strip() in ("1", "是", "y", "Y") else 0,
                  row.get("sales_volume_band"), row.get("sales_period")))
        report["pricing"] += 1

    return report


def _rows(wb, sheet: str, cols: list, report: dict):
    if sheet not in wb.sheetnames:
        report["warnings"].append(f"缺少工作表「{sheet}」，已跳过")
        return
    ws = wb[sheet]
    keys = [c[0] for c in cols]
    labels = [c[1] for c in cols]
    header_row = None
    for r in range(1, min(ws.max_row, 8) + 1):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, len(cols) + 1)]
        if vals and vals[0] == labels[0]:
            header_row = r
            break
    if header_row is None:
        report["errors"].append(f"工作表「{sheet}」找不到表头行（第一列应为「{labels[0]}」）")
        return
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(cols) + 1)]
        if all(v in (None, "") for v in vals):
            continue
        row = dict(zip(keys, vals))
        # 跳过模板自带的示例行
        if str(row.get(keys[0]) or "").strip() == cols[0][2]:
            continue
        yield {k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()}


def _int(v):
    try:
        return int(float(str(v).strip())) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _float(v):
    try:
        return float(str(v).replace(",", "").strip()) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


def _date(v):
    if v in (None, ""):
        return None
    s = str(v)[:10]
    import re
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else None


# ---------------------------------------------------------------- CSV 清单导入

def import_product_list(path: str | Path | None = None) -> dict:
    """导入产品清单 CSV（product_line / product_family / product_series /
    product / category_code）。

    ★ 唯一键是「营销名 + 内部代号」，不是营销名。
      清单里有 7 组同名不同代号的产品（如 WATCH GT 6 有 Aris 和 Kelo 两个平台），
      只用营销名会把它们合并成一个，丢掉一半产品。
    """
    import csv as _csv
    from . import config as _config

    path = Path(path) if path else (_config.CONFIG_DIR / "my_products.csv")
    if not path.exists():
        return {"error": f"找不到清单文件 {path}"}

    valid_cats = {r["code"] for r in db.q("SELECT code FROM category")}
    report = {"total": 0, "created": 0, "updated": 0, "errors": []}

    # utf-8-sig 吃掉 Excel 存 CSV 时加的 BOM
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        for row in _csv.DictReader(f):
            name = (row.get("product") or "").strip()
            code = (row.get("product_series") or "").strip()
            cat = (row.get("category_code") or "").strip().lower()
            if not name:
                continue
            report["total"] += 1
            if cat not in valid_cats:
                report["errors"].append(f"「{name}」产业代码 {cat} 无效，已跳过")
                continue

            existing = db.q1("""SELECT id FROM my_product
                                WHERE marketing_name=? AND COALESCE(internal_code,'')=?""",
                             (name, code))
            with db.tx() as conn:
                if existing:
                    conn.execute("""UPDATE my_product SET category_code=?, series=?,
                                    updated_at=datetime('now') WHERE id=?""",
                                 (cat, (row.get("product_family") or "").strip(),
                                  existing["id"]))
                    report["updated"] += 1
                else:
                    conn.execute("""
                        INSERT INTO my_product(marketing_name,internal_code,
                            category_code,series,status)
                        VALUES(?,?,?,?,'active')
                    """, (name, code or None, cat,
                          (row.get("product_family") or "").strip()))
                    report["created"] += 1
    return report


# ---------------------------------------------------------------- 查询

def list_products() -> list[dict]:
    rows = db.q("""
        SELECT mp.*, c.name_zh AS category_name, c.icon,
               (SELECT COUNT(*) FROM my_sku WHERE product_id=mp.id) AS sku_count,
               (SELECT COUNT(*) FROM my_pricing WHERE product_id=mp.id AND on_sale=1)
                 AS country_count,
               (SELECT COUNT(*) FROM competitor_match WHERE my_product_id=mp.id
                  AND is_excluded=0) AS match_count
        FROM my_product mp LEFT JOIN category c ON c.code=mp.category_code
        ORDER BY c.sort_order, mp.marketing_name
    """)
    for r in rows:
        try:
            r["box_contents"] = json.loads(r.get("box_contents") or "[]")
        except Exception:  # noqa: BLE001
            r["box_contents"] = []
    return rows


def product_detail(pid: int) -> dict | None:
    p = db.q1("SELECT * FROM my_product WHERE id=?", (pid,))
    if not p:
        return None
    try:
        p["box_contents"] = json.loads(p.get("box_contents") or "[]")
    except Exception:  # noqa: BLE001
        p["box_contents"] = []
    p["skus"] = db.q("SELECT * FROM my_sku WHERE product_id=? ORDER BY rom_gb", (pid,))
    p["pricing"] = db.q("""
        SELECT mp.*, co.name_zh AS country_name, co.sort_order
        FROM my_pricing mp JOIN country co ON co.code=mp.country_code
        WHERE mp.product_id=? ORDER BY co.sort_order
    """, (pid,))
    return p
